"""
Resume Data Extractor → Excel
Reads all resume PDFs/DOCX from the uploads folder,
sends them to Groq API (Llama 3.3) to extract important details,
and saves everything into a styled Excel file with serial numbers.

Usage:
    python resume_to_excel.py
    python resume_to_excel.py --folder path/to/resumes
    python resume_to_excel.py --output my_report.xlsx
"""

import os
import sys
import io
import json
import time
import argparse
from datetime import datetime

# Fix Windows console encoding for Unicode/emoji
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ─── PDF & DOCX parsing ────────────────────────────────────────────
from PyPDF2 import PdfReader
from docx import Document

# ─── Groq API ──────────────────────────────────────────────────────
from groq import Groq

# ─── Excel output ──────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  CONFIG
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
GROQ_API_KEY = "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"
GROQ_MODEL = "llama-3.3-70b-versatile"
DEFAULT_FOLDER = os.path.join(os.path.dirname(__file__), "server", "uploads")
SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".doc", ".txt"}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  RESUME PARSING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def extract_text_from_pdf(filepath):
    """Extract text from a PDF file."""
    reader = PdfReader(filepath)
    text = ""
    for page in reader.pages:
        page_text = page.extract_text()
        if page_text:
            text += page_text + "\n"
    return text.strip()


def extract_text_from_docx(filepath):
    """Extract text from a DOCX file."""
    doc = Document(filepath)
    return "\n".join(para.text for para in doc.paragraphs).strip()


def extract_text_from_txt(filepath):
    """Extract text from a plain text file."""
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        return f.read().strip()


def extract_text(filepath):
    """Extract text from any supported resume file."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".pdf":
        return extract_text_from_pdf(filepath)
    elif ext in (".docx", ".doc"):
        return extract_text_from_docx(filepath)
    elif ext == ".txt":
        return extract_text_from_txt(filepath)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  GROQ API — Extract Important Resume Details
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SYSTEM_PROMPT = """You are an expert resume parser. Extract ONLY the important details from the resume text.
Return valid JSON only — no explanation, no markdown fences, no extra text.

RULES:
- Extract ONLY the fields listed below
- If a field is not found, use null for strings or [] for arrays
- Keep it concise — only important information
- Phone numbers should include country code if present
- Skills should be individual skill names as simple strings
- Projects: just title and technologies used

Return this EXACT JSON structure:
{
  "name": "<full name or null>",
  "email": "<primary email or null>",
  "phone": "<primary phone or null>",
  "degree": "<highest degree like B.Tech, BCA, MCA, etc. or null>",
  "stream": "<branch like CSE, IT, ECE, etc. or null>",
  "cgpa": "<CGPA or percentage or null>",
  "tenth_marks": "<10th marks/percentage or null>",
  "twelfth_marks": "<12th marks/percentage or null>",
  "github": "<GitHub URL or null>",
  "linkedin": "<LinkedIn URL or null>",
  "portfolio": "<Portfolio/website URL or null>",
  "skills": ["<skill1>", "<skill2>", "<skill3>"],
  "projects": ["<project1 title>", "<project2 title>"],
  "certifications": ["<cert1>", "<cert2>"],
  "experience": "<total experience summary or null>",
  "languages": ["<lang1>", "<lang2>"]
}"""


def extract_with_groq(resume_text):
    """Send resume text to Groq API and get structured data back."""
    client = Groq(api_key=GROQ_API_KEY)

    # Truncate to avoid token limits
    truncated = resume_text[:6000]

    try:
        response = client.chat.completions.create(
            model=GROQ_MODEL,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {
                    "role": "user",
                    "content": f"Extract the important details from this resume:\n\n{truncated}",
                },
            ],
            temperature=0.1,
            max_tokens=1024,
            response_format={"type": "json_object"},
        )

        raw = response.choices[0].message.content.strip()

        # Parse JSON
        data = json.loads(raw)
        return data

    except json.JSONDecodeError as e:
        print(f"  ⚠ JSON parse error: {e}")
        return None
    except Exception as e:
        print(f"  ⚠ Groq API error: {e}")
        return None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  EXCEL GENERATION — Styled & Professional
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def arr_to_str(value):
    """Convert an array or value to a semicolon-separated string."""
    if isinstance(value, list):
        return "; ".join(str(v) for v in value if v)
    return str(value) if value else ""


def create_excel(all_data, output_path):
    """Create a beautifully styled Excel file from extracted resume data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Resume Data"

    # ─── Define Styles ──────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    title_font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    title_alignment = Alignment(horizontal="center", vertical="center")

    data_font = Font(name="Calibri", size=11)
    data_alignment = Alignment(vertical="top", wrap_text=True)

    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    serial_font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    serial_alignment = Alignment(horizontal="center", vertical="top")

    # ─── Title Row ──────────────────────────────────────────────────
    ws.merge_cells("A1:Q1")
    title_cell = ws["A1"]
    title_cell.value = f"📋 Resume Data Report — Generated {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    ws.row_dimensions[1].height = 40

    # ─── Empty spacer row ──────────────────────────────────────────
    ws.row_dimensions[2].height = 10

    # ─── Headers ────────────────────────────────────────────────────
    headers = [
        ("S.No", 6),
        ("File Name", 25),
        ("Name", 20),
        ("Email", 28),
        ("Phone", 18),
        ("Degree", 14),
        ("Stream", 14),
        ("CGPA/%", 10),
        ("10th Marks", 12),
        ("12th Marks", 12),
        ("GitHub", 30),
        ("LinkedIn", 30),
        ("Portfolio", 28),
        ("Skills", 45),
        ("Projects", 35),
        ("Certifications", 35),
        ("Experience", 25),
        ("Languages", 20),
    ]

    header_row = 3
    for col_idx, (header_name, width) in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[header_row].height = 30

    # ─── Data Rows ─────────────────────────────────────────────────
    for idx, item in enumerate(all_data, start=1):
        row_num = header_row + idx
        data = item.get("data") or {}
        fill = even_fill if idx % 2 == 0 else odd_fill

        row_values = [
            idx,                                          # S.No
            item.get("filename", ""),                     # File Name
            data.get("name", ""),                         # Name
            data.get("email", ""),                        # Email
            data.get("phone", ""),                        # Phone
            data.get("degree", ""),                       # Degree
            data.get("stream", ""),                       # Stream
            data.get("cgpa", ""),                         # CGPA
            data.get("tenth_marks", ""),                  # 10th
            data.get("twelfth_marks", ""),                # 12th
            data.get("github", ""),                       # GitHub
            data.get("linkedin", ""),                     # LinkedIn
            data.get("portfolio", ""),                    # Portfolio
            arr_to_str(data.get("skills", [])),           # Skills
            arr_to_str(data.get("projects", [])),         # Projects
            arr_to_str(data.get("certifications", [])),   # Certifications
            data.get("experience", ""),                   # Experience
            arr_to_str(data.get("languages", [])),        # Languages
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value or "")
            cell.border = thin_border
            cell.fill = fill

            if col_idx == 1:  # Serial number column
                cell.font = serial_font
                cell.alignment = serial_alignment
            else:
                cell.font = data_font
                cell.alignment = data_alignment

        ws.row_dimensions[row_num].height = 28

    # ─── Freeze panes (header stays visible) ───────────────────────
    ws.freeze_panes = "A4"

    # ─── Auto-filter ───────────────────────────────────────────────
    last_col = get_column_letter(len(headers))
    last_row = header_row + len(all_data)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"

    # ─── Summary Sheet ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "📊 Resume Processing Summary"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")

    summary_data = [
        ("Total Resumes Processed", len(all_data)),
        ("Successful Extractions", sum(1 for d in all_data if d.get("data"))),
        ("Failed Extractions", sum(1 for d in all_data if not d.get("data"))),
        ("Generated On", datetime.now().strftime("%d %b %Y, %I:%M %p")),
        ("AI Model Used", GROQ_MODEL),
    ]

    for i, (label, value) in enumerate(summary_data, start=3):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Calibri", bold=True, size=11)
        ws2.cell(row=i, column=2, value=value).font = Font(name="Calibri", size=11)
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 40

    # ─── Save ──────────────────────────────────────────────────────
    wb.save(output_path)
    return output_path


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  MAIN
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def main():
    parser = argparse.ArgumentParser(description="Extract resume data → Excel using Groq AI")
    parser.add_argument("--folder", default=DEFAULT_FOLDER, help="Path to folder containing resumes")
    parser.add_argument("--output", default=None, help="Output Excel file path")
    args = parser.parse_args()

    folder = args.folder
    output = args.output or os.path.join(
        os.path.dirname(__file__),
        f"resumes_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )

    if not os.path.isdir(folder):
        print(f"❌ Folder not found: {folder}")
        sys.exit(1)

    # Find all resume files
    files = [
        f
        for f in os.listdir(folder)
        if os.path.splitext(f)[1].lower() in SUPPORTED_EXTENSIONS
    ]

    if not files:
        print(f"❌ No resume files found in: {folder}")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  📄 Resume Data Extractor → Excel")
    print(f"  📁 Folder: {folder}")
    print(f"  📝 Found {len(files)} resume(s)")
    print(f"  🤖 Using: Groq API ({GROQ_MODEL})")
    print(f"{'='*60}\n")

    all_data = []

    for i, filename in enumerate(files, start=1):
        filepath = os.path.join(folder, filename)
        print(f"  [{i}/{len(files)}] Processing: {filename}")

        try:
            # Step 1: Extract text from file
            text = extract_text(filepath)
            if not text or len(text) < 50:
                print(f"    ⚠ Too little text extracted, skipping...")
                all_data.append({"filename": filename, "data": None})
                continue

            print(f"    ✓ Text extracted ({len(text)} chars)")

            # Step 2: Send to Groq API
            print(f"    → Sending to Groq AI...")
            data = extract_with_groq(text)

            if data:
                name = data.get("name", "Unknown")
                skills_count = len(data.get("skills", []))
                print(f"    ✓ Extracted: {name} | {skills_count} skills found")
                all_data.append({"filename": filename, "data": data})
            else:
                print(f"    ⚠ Failed to extract data")
                all_data.append({"filename": filename, "data": None})

            # Small delay to respect rate limits
            if i < len(files):
                time.sleep(1)

        except Exception as e:
            print(f"    ❌ Error: {e}")
            all_data.append({"filename": filename, "data": None})

    # Step 3: Generate Excel
    print(f"\n  📊 Generating Excel file...")
    output_path = create_excel(all_data, output)

    success = sum(1 for d in all_data if d.get("data"))
    print(f"\n{'='*60}")
    print(f"  ✅ Done! {success}/{len(files)} resumes processed successfully")
    print(f"  📁 Excel saved: {output_path}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
